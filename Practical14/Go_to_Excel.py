import xml.sax
import openpyxl

class GOHandler(xml.sax.ContentHandler):
    def __init__(self):
        self.current_element = ""
        self.go_id = ""
        self.term_name = ""
        self.definition = ""
        self.go_parents = {}
        # Starting row index in the spreadsheet
        self.row_index = 2
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet["A1"] = "GO ID"
        self.sheet["B1"] = "Term Name"
        self.sheet["C1"] = "Definition"
        self.sheet["D1"] = "Child Node Count"

    def startElement(self, name, attrs):
        self.current_element = name

    def endElement(self, name):
        if name == "term":
            if "autophagosome" in self.definition:
                self.sheet.cell(row=self.row_index, column=1).value = self.go_id
                self.sheet.cell(row=self.row_index, column=2).value = self.term_name
                self.sheet.cell(row=self.row_index, column=3).value = self.definition
                self.row_index += 1
            # Reset the variables for the next term
            self.go_id = ""
            self.term_name = ""
            self.definition = ""
        self.current_element = ""

    def characters(self, content):
        if self.current_element == "id":
            self.go_id = content
        elif self.current_element == "name":
            self.term_name = content
        elif self.current_element == "defstr":
            self.definition = content
        elif self.current_element == "is_a":
            parent_id = content
            self.go_parents.setdefault(parent_id, []).append(self.go_id)

    def get_child_node_count(self, go_id):
        if go_id not in self.go_parents:
            return 0
        count = len(self.go_parents[go_id])
        for parent_id in self.go_parents[go_id]:
            count += self.get_child_node_count(parent_id)
        return count

    def write_excel_file(self, filename):
        for row in self.sheet.iter_rows(min_row=2, max_row=self.row_index-1):
            go_id = row[0].value
            child_count = self.get_child_node_count(go_id)
            self.sheet.cell(row=row[0].row, column=4).value = child_count
        self.workbook.save(filename)

# Create an instance of the SAX parser
parser = xml.sax.make_parser()
# Create an instance of the content handler
handler = GOHandler()
# Set the content handler to the parser
parser.setContentHandler(handler)
# Parse the XML file
parser.parse("go_obo.xml")
# Save the workbook to an Excel file
handler.write_excel_file("autophagosome.xlsx")
